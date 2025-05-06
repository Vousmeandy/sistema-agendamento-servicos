import sqlite3
import openpyxl

class Database:
    def __init__(self, db_file, excel_file):
        self.db_file = db_file
        self.excel_file = excel_file

    def criar_tabelas(self):
        conn = self.obter_conexao()
        cursor = conn.cursor()

        # Criação da tabela de clientes
        cursor.execute('''CREATE TABLE IF NOT EXISTS clientes
                          (id_cliente INTEGER PRIMARY KEY AUTOINCREMENT, cpf TEXT, nome TEXT, data_nascimento TEXT)''')

        # Criação da tabela de funcionários
        cursor.execute('''CREATE TABLE IF NOT EXISTS funcionarios
                          (id_funcionario INTEGER PRIMARY KEY AUTOINCREMENT, cpf TEXT, nome TEXT, data_nascimento TEXT, salario REAL)''')

        # Criação da tabela de serviços
        cursor.execute('''CREATE TABLE IF NOT EXISTS servicos
                          (id_servico INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT, preco REAL, comissao REAL)''')

        conn.commit()
        conn.close()

    @staticmethod
    def executar_query(query, parametros=None):
        conn = sqlite3.connect(Database.db_file)
        cursor = conn.cursor()

        if parametros:
            cursor.execute(query, parametros)
        else:
            cursor.execute(query)

        conn.commit()
        conn.close()
    def obter_conexao(self):
        return sqlite3.connect(self.db_file)

    def exportar_para_excel(self):
        conn = self.obter_conexao()
        cursor = conn.cursor()

        # Consulta dos clientes
        cursor.execute("SELECT * FROM clientes")
        clientes = cursor.fetchall()

        # Consulta dos funcionários
        cursor.execute("SELECT * FROM funcionarios")
        funcionarios = cursor.fetchall()

        conn.close()

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Cabeçalhos das colunas
        sheet.cell(row=1, column=1, value="ID Cliente")
        sheet.cell(row=1, column=2, value="CPF")
        sheet.cell(row=1, column=3, value="Nome")
        sheet.cell(row=1, column=4, value="Data de Nascimento")
        sheet.cell(row=1, column=5, value="Salário")

        # Preenchimento dos dados dos clientes
        for row_index, cliente in enumerate(clientes, start=2):
            sheet.cell(row=row_index, column=1, value=cliente[0])
            sheet.cell(row=row_index, column=2, value=cliente[1])
            sheet.cell(row=row_index, column=3, value=cliente[2])
            sheet.cell(row=row_index, column=4, value=cliente[3])

        # Preenchimento dos dados dos funcionários
        for row_index, funcionario in enumerate(funcionarios, start=2):
            sheet.cell(row=row_index, column=2, value=funcionario[1])
            sheet.cell(row=row_index, column=3, value=funcionario[2])
            sheet.cell(row=row_index, column=4, value=funcionario[3])
            sheet.cell(row=row_index, column=5, value=funcionario[4])

        workbook.save(self.excel_file)

database = Database("clientes.db", "clientes.xlsx")
database.criar_tabelas()
database.exportar_para_excel()

